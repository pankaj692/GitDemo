import openpyxl

book = openpyxl.load_workbook("D:\\pythonProject\\PythonSelFramework\\PythonDemo.xlsx")

# Read in sheet
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

print(sheet['A5'].value)

# write in sheet
sheet.cell(row=2, column=2).value = "Pankaj"
print(sheet.cell(row=2, column=2).value)

# for max row and column
print(sheet.max_row)
print(sheet.max_column)

# print every value
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TestCase2":

        for j in range(2, sheet.max_column+1):

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            # print(sheet.cell(row=i, column=j).value)

print(Dict)